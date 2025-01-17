import datetime
from datetime import datetime, timedelta
from robot.api.deco import keyword
from openpyxl import load_workbook
import json
import random
import string
import re
import pandas as pd
import os
import csv
import shutil
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
import ast
import random
import calendar
from datetime import datetime, timedelta

@keyword
def fetch_testdata_by_id(file_path, sheet_name, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except NameError:
            pass
    return None


@keyword
def fetch_user_login_data(json_file, user_type, user_id):
    try:
        with open(json_file, 'r') as file:
            data = json.load(file)
            users = data.get('users', {}).get(user_type, [])
            for user in users:
                try:
                    if user.get('user_id') == user_id:
                        user_data = {'Username': user.get('user_name'), 'Password': user.get('password')}
                        return user_data
                except (KeyError, ValueError):
                    print("Error processing user:", user)
                    continue
            raise ValueError("User not found or incorrect user ID")
        print("User not found.")
        return None

    except (FileNotFoundError, json.JSONDecodeError, KeyError) as e:
        raise type(e)(str(e) + f" in JSON file '{json_file}'")

@keyword
def no_of_months_ahead(a):
    current_date = datetime.now()
    a = int(a)
    new_month = current_date.month + a
    # Calculate the new year by adding 1 to the current year if the new month is greater than 12
    new_year = current_date.year + (new_month - 1) // 12

    # Calculate the new month (taking into account that 12 is December)
    new_month = (new_month - 1) % 12 + 1

    # Calculate the day of the new date (taking care of leap years and different month lengths)
    new_day = min(current_date.day, [31,
                                     29 if new_year % 4 == 0 and (new_year % 100 != 0 or new_year % 400 == 0) else 28,
                                     31, 30, 31, 30, 31, 31, 30, 31, 30, 31][new_month - 1])

    # Create the new date
    new_date = datetime(new_year, new_month, new_day)

    print("Current Date:", current_date.strftime("%d-%m-%Y"))
    print("New Date after adding 3 months:", new_date.strftime("%Y-%m-%d"))
    str = new_date.strftime("%d-%m-%Y")
    print(current_date.strftime("%B"))
    monthnew = (new_date.strftime("%B"))
    daynew = (new_date.strftime("%d"))
    tmp = ""
    print("old ", daynew[1])
    if (daynew[0] == "0"):
        daynew = daynew[1]
    print("removed zero ", daynew)

    Yearnew = (new_date.strftime("%Y"))
    print(daynew, " ", Yearnew, " ", monthnew)

    mainlist = [daynew, monthnew, Yearnew]
    return mainlist
